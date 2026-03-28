import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from typing import Tuple, Optional

EXPECTED_COLUMNS_SET = {"почта", "имя", "пол", "дата", "время", "место", "задание"}
EXPECTED_COLUMNS_ORDER = ["почта", "имя", "пол", "дата", "время", "место", "задание"]


class ExcelParser:
    @staticmethod
    def load_excel(file_path: str) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
        try:
            df = pd.read_excel(file_path, dtype=str).fillna('')
        except Exception as e:
            return None, f"Ошибка чтения файла: {e}"

        df.columns = df.columns.str.strip().str.lower()
        cols_set = set(df.columns)
        extra = cols_set - EXPECTED_COLUMNS_SET
        if extra:
            return None, f"Найдены лишние столбцы: {', '.join(extra)}"

        missing = EXPECTED_COLUMNS_SET - cols_set
        if missing:
            return df, f"Отсутствуют столбцы: {', '.join(missing)}. Они могут понадобиться, если будут использованы соответствующие метки."
        return df, None

    @staticmethod
    def generate_template(save_path: str):
        """Создаёт шаблон Excel с текстовым форматом ячеек"""
        df = pd.DataFrame(columns=EXPECTED_COLUMNS_ORDER)
        # Сохраняем через openpyxl
        df.to_excel(save_path, index=False, engine='openpyxl')
        # Открываем книгу и устанавливаем текстовый формат для всех ячеек, кроме заголовков
        wb = load_workbook(save_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):  # начиная со второй строки (данные)
            for cell in row:
                cell.number_format = '@'  # текстовый формат
        wb.save(save_path)